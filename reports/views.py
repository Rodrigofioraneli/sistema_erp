from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django import forms
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from django.db import transaction
from django.db.utils import OperationalError, ProgrammingError
from django.db.models import Sum, Count, F, Avg, Q, Case, When, Value
from django.db.models.functions import TruncMonth, ExtractHour
from django.utils import timezone
from django.conf import settings
from datetime import date, timedelta
from django.core.serializers.json import DjangoJSONEncoder
from django.utils.encoding import escape_uri_path
import json
import unicodedata
import csv
import os
from io import TextIOWrapper
import calendar
import re
from datetime import datetime

from django.contrib.auth.models import User
# Tratamento de erro para bibliotecas opcionais (evita que o servidor pare se não estiverem instaladas)
try:
    import openpyxl
except ImportError:
    openpyxl = None

from sales.models import Sale, SaleItem, AuditLog
from products.models import Product, Brand, OlfactoryFamily, StockMovement, Category, Supplier, ProductComponent
from customers.models import Customer
# from finance.models import Expense  <-- Removido, agora importamos do local correto
from .models import CompanySettings, PaymentMethod, Expense

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
except ImportError:
    canvas = None

from sales.decorators import admin_required  # Importando nosso protetor
from django.urls import reverse

def normalize_str(s):
    """Remove acentos e coloca em minúsculo para busca robusta"""
    if s is None: return ''
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn').lower()

def clean_br_decimal(val_str):
    """Helper para converter strings de moeda (pt-BR ou en-US) para Decimal"""
    if not val_str: return Decimal('0')
    s = str(val_str).strip().replace('R$', '').replace(' ', '').replace('\xa0', '')
    if not s: return Decimal('0')
    
    # Força padrão BR: remove pontos (milhar) e troca vírgula por ponto (decimal)
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        s = s.replace('.', '')
        
    try: return Decimal(s)
    except: return Decimal('0')

# --- Configuração da Empresa (Home) ---

class CompanySettingsForm(forms.ModelForm):
    class Meta:
        model = CompanySettings
        fields = ['name', 'cnpj', 'state_registration', 'address', 'phone', 'email', 'website', 'logo',
                  'primary_color', 'secondary_color', 'background_color', 'font_family', 'font_size']
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for field in self.fields.values():
            field.widget.attrs['class'] = 'form-control'
            if isinstance(field.widget, forms.FileInput):
                field.widget.attrs['class'] = 'form-control'

@admin_required
@login_required
def settings_dashboard(request):
    """
    Nova aba de Configurações da Loja.
    Centraliza edição de Empresa, Categorias, Fornecedores, etc.
    """
    # 1. Processamento do Formulário da Empresa
    try:
        company = CompanySettings.objects.first()
        if not company:
            company = CompanySettings.objects.create()
    except (OperationalError, ProgrammingError):
        company = None

    # --- Lógica para Adicionar/Remover Itens (Marcas, Categorias, etc.) ---
    if request.method == 'POST' and 'item_action' in request.POST:
        action = request.POST.get('item_action')
        model_type = request.POST.get('model_type')
        item_name = request.POST.get('item_name')
        item_id = request.POST.get('item_id')

        model_map = {
            'brand': Brand,
            'category': Category,
            'family': OlfactoryFamily,
            'supplier': Supplier,
            'payment_method': PaymentMethod
        }

        ModelClass = model_map.get(model_type)
        if ModelClass:
            if action == 'add' and item_name:
                defaults = {}
                if model_type == 'supplier':
                    defaults = {'cnpj_cpf': '', 'phone': '', 'email': ''}
                
                ModelClass.objects.get_or_create(name=item_name.strip(), defaults=defaults)
                messages.success(request, f'Item adicionado com sucesso!')
            elif action == 'delete' and item_id:
                try:
                    ModelClass.objects.filter(pk=item_id).delete()
                    messages.success(request, 'Item removido com sucesso!')
                except Exception as e:
                    messages.error(request, f'Erro ao remover: {str(e)}')
        
        # Redireciona mantendo a aba ativa
        target_tab = 'others' # Todos os itens auxiliares ficam na aba 'others'
        return redirect(f"{reverse('settings_dashboard')}?tab={target_tab}")

    # --- Lógica para Gestão de Funcionários (NOVO) ---
    if request.method == 'POST' and 'employee_action' in request.POST:
        action = request.POST.get('employee_action')
        user_id = request.POST.get('user_id')
        
        if action == 'add':
            username = request.POST.get('username')
            password = request.POST.get('password')
            is_admin = request.POST.get('is_admin') == 'on'
            
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Erro: Já existe um usuário com este login.')
            else:
                # Cria usuário (Superusuário se marcado como Admin, caso contrário normal)
                if is_admin:
                    User.objects.create_superuser(username=username, email='', password=password)
                else:
                    User.objects.create_user(username=username, email='', password=password)
                messages.success(request, f'Funcionário "{username}" cadastrado com sucesso!')

        elif action == 'delete':
            if str(request.user.id) == str(user_id):
                messages.error(request, 'Você não pode excluir a si mesmo!')
            else:
                User.objects.filter(pk=user_id).delete()
                messages.success(request, 'Funcionário removido.')
        return redirect(f"{reverse('settings_dashboard')}?tab=employees")

    # --- Lógica para Salvar Tema (Aparência) ---
    if request.method == 'POST' and 'theme_form' in request.POST:
        form = CompanySettingsForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aparência do sistema atualizada com sucesso!')
            return redirect(f"{reverse('settings_dashboard')}?tab=appearance")

    if request.method == 'POST' and 'company_form' in request.POST:
        form = CompanySettingsForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dados da empresa atualizados!')
            return redirect(f"{reverse('settings_dashboard')}?tab=company")
    else:
        form = CompanySettingsForm(instance=company)

    # Define a aba ativa com base no parâmetro GET (padrão: company)
    active_tab = request.GET.get('tab', 'company')

    # 2. Dados para Listagem nas Abas de Configuração
    context = {
        'company_form': form,
        'company': company,
        'categories': Category.objects.all(),
        'suppliers': Supplier.objects.all(),
        'users_list': User.objects.all().order_by('username'), # Lista de funcionários
        'payment_methods': PaymentMethod.objects.all(),
        'olfactory_families': OlfactoryFamily.objects.all(),
        'brands': Brand.objects.all(),
        'active_tab': active_tab,
    }
    return render(request, 'reports/settings.html', context)

def home_view(request):
    """
    Antiga 'company_settings'. Agora é apenas a Home com Logo.
    """
    if not request.user.is_authenticated:
        return render(request, 'home.html')
    
    try:
        company = CompanySettings.objects.first()
    except (OperationalError, ProgrammingError):
        company = None
    
    return render(request, 'home.html', {'company': company})

# Alias para manter compatibilidade com config/urls.py que busca a view antiga
company_settings = home_view

def _calculate_sales_metrics(start_date, end_date):
    """
    Função Auxiliar (Privada): Calcula todos os KPIs (Indicadores) financeiros
    para um intervalo de datas específico. Retorna um dicionário com os valores.
    """
    # Filtra vendas finalizadas
    sales_qs = Sale.objects.filter(status='completed')
    if start_date:
        sales_qs = sales_qs.filter(created_at__date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(created_at__date__lte=end_date)
    
    expenses_qs = Expense.objects.all()
    if start_date:
        expenses_qs = expenses_qs.filter(date__gte=start_date)
    if end_date:
        expenses_qs = expenses_qs.filter(date__lte=end_date)

    # Soma o total das vendas
    total_revenue = sales_qs.aggregate(Sum('total'))['total__sum'] or 0
    total_count = sales_qs.count()
    avg_ticket = (total_revenue / total_count) if total_count > 0 else 0

    # Calcula Custo da Mercadoria Vendida (CMV)
    total_cost = SaleItem.objects.filter(sale__in=sales_qs).aggregate(
        cost=Sum(F('product__cost_price') * F('quantity'))
    )['cost'] or 0

    total_expenses = expenses_qs.aggregate(Sum('amount'))['amount__sum'] or 0
    gross_profit = total_revenue - total_cost
    net_profit = gross_profit - total_expenses
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

    return {
        'total_revenue': float(total_revenue),
        'total_count': int(total_count),
        'avg_ticket': float(avg_ticket),
        'total_cost': float(total_cost),
        'gross_profit': float(gross_profit),
        'total_expenses': float(total_expenses),
        'net_profit': float(net_profit),
        'profit_margin': float(profit_margin)
    }

@admin_required
@login_required
def reports_dashboard(request):
    """
    View Principal do Dashboard.
    Esta view centraliza a inteligência de negócios do ERP.
    Calcula KPIs, prepara dados para Chart.js e filtra resultados por data.
    """
    # Filtros básicos de data (podem ser expandidos depois)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    period = request.GET.get('period')
    
    # [MELHORIA] Adiciona filtros para o período de comparação
    compare_start_date_str = request.GET.get('compare_start_date')
    compare_end_date_str = request.GET.get('compare_end_date')

    # Data de hoje para referências
    today = timezone.now().date()
    
    # FIX: Se nenhum filtro for passado, define o mês atual como padrão para evitar mostrar "Tudo"
    if not start_date_str and not end_date_str and not period:
        start_date_str = today.replace(day=1).strftime('%Y-%m-%d')
        end_date_str = today.strftime('%Y-%m-%d')
        period = 'month'

    # Lógica de atalhos de data (Dia, Mês, Ano)
    if period:
        if period == 'today':
            start_date_str = today.strftime('%Y-%m-%d')
            end_date_str = today.strftime('%Y-%m-%d')
        elif period == 'month':
            start_date_str = today.replace(day=1).strftime('%Y-%m-%d')
            end_date_str = today.strftime('%Y-%m-%d')
        elif period == 'year':
            start_date_str = today.replace(month=1, day=1).strftime('%Y-%m-%d')
            end_date_str = today.strftime('%Y-%m-%d')
    
    # 1. Métricas de Vendas e Financeiras (período principal)
    sales_metrics = _calculate_sales_metrics(start_date_str, end_date_str)

    # [MELHORIA] Lógica de Comparação de Períodos
    if compare_start_date_str and compare_end_date_str:
        comparison_metrics = _calculate_sales_metrics(compare_start_date_str, compare_end_date_str)
        
        # Calcula a variação percentual para cada KPI
        for key in ['total_revenue', 'total_count', 'avg_ticket', 'net_profit', 'total_expenses']:
            current_val = sales_metrics.get(key, 0)
            previous_val = comparison_metrics.get(key, 0)
            
            if previous_val != 0:
                change = ((current_val - previous_val) / abs(previous_val)) * 100
            elif current_val > 0:
                change = 100.0 # Crescimento "infinito" se o anterior era zero
            else:
                change = 0.0
            
            # Inverte a lógica para despesas: uma queda é um resultado positivo
            if key == 'total_expenses':
                change *= -1

            sales_metrics[f'{key}_change'] = change

    # Recria o queryset principal para os relatórios de "Top Produtos", etc.
    sales_qs = Sale.objects.filter(status='completed')
    if start_date_str:
        sales_qs = sales_qs.filter(created_at__date__gte=start_date_str)
    if end_date_str:
        sales_qs = sales_qs.filter(created_at__date__lte=end_date_str)
    
    # 1.2 Evolução Mensal (Vendas e Lucros)
    # Gera loop preciso mês a mês para popular o gráfico, garantindo que despesas entrem
    # mesmo em meses sem vendas.
    monthly_profit = []
    
    if start_date_str and end_date_str:
        try:
            # Converte para objetos date
            dt_start = datetime.strptime(start_date_str, '%Y-%m-%d').date().replace(day=1)
            dt_end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            # Loop iterando pelo primeiro dia de cada mês
            current_dt = dt_start
            
            while current_dt <= dt_end.replace(day=1):
                year = current_dt.year
                month = current_dt.month
                
                # 1. Receita do Mês (Vendas Finalizadas)
                month_revenue = Sale.objects.filter(
                    status='completed',
                    created_at__year=year, 
                    created_at__month=month
                ).aggregate(total=Sum('total'))['total'] or 0
                
                # 2. Custo do Produto (CMV) do Mês
                month_cost = SaleItem.objects.filter(
                    sale__status='completed', 
                    sale__created_at__year=year, 
                    sale__created_at__month=month
                ).aggregate(cost=Sum(F('product__cost_price') * F('quantity')))['cost'] or 0
                
                # 3. Despesas Operacionais do Mês (Considera parcelas que caem neste mês)
                month_expenses = Expense.objects.filter(
                    date__year=year, date__month=month
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

                # Apenas adiciona ao gráfico se houver alguma movimentação financeira
                if month_revenue > 0 or month_expenses > 0:
                    monthly_profit.append({
                        'month': current_dt.strftime('%Y-%m'),
                        'revenue': float(month_revenue),
                        'cost': float(month_cost),
                        'expenses': float(month_expenses),
                        # Lucro Líquido = Receita - Custo Produto - Despesas
                        'profit': float(month_revenue - month_cost - month_expenses) 
                    })
                
                # Avança para o próximo mês
                days_in_month = calendar.monthrange(current_dt.year, current_dt.month)[1]
                current_dt = current_dt + timedelta(days=days_in_month)
                current_dt = current_dt.replace(day=1)
        except ValueError:
            pass # Datas inválidas

    # 1.3 Vendas por Forma de Pagamento
    payment_stats_qs = sales_qs.values('payment_method')\
        .annotate(total=Sum('total'), count=Count('id'))\
        .order_by('-total')
    
    payment_stats = list(payment_stats_qs)

    # 1.4 Horários de Pico (Operacional)
    peak_hours_qs = sales_qs.annotate(hour=ExtractHour('created_at'))\
        .values('hour')\
        .annotate(count=Count('id'))\
        .order_by('hour')
    
    peak_hours = [{'hour': item['hour'], 'count': item['count']} for item in peak_hours_qs]

    # 2. Top Produtos
    top_products = SaleItem.objects.filter(sale__in=sales_qs)\
        .values('product__name')\
        .annotate(total_qty=Sum('quantity'))\
        .order_by('-total_qty')[:5]
        
    # 3. Métricas de Estoque
    stock_metrics = Product.objects.aggregate(
        total_value=Sum(F('stock_quantity') * F('cost_price')),
        total_items=Sum('stock_quantity')
    )
    
    low_stock_count = Product.objects.filter(stock_quantity__lte=5).count()
    
    # Lista de produtos com estoque baixo para exibir na tabela
    low_stock_products = Product.objects.filter(stock_quantity__lte=5)\
        .values('name', 'stock_quantity', 'selling_price', 'image', 'image_url')\
        .order_by('stock_quantity')[:10]
    
    # Lista de vendas recentes para exibição em tabela
    recent_sales = sales_qs.select_related('customer').order_by('-created_at')[:20]

    # 4. Visualização de Relatório na Tela (Se solicitado)
    report_type = request.GET.get('report_type')
    report_data = None
    if report_type:
        # Gera os dados do relatório escolhido para exibir na tela
        r_title, r_headers, r_data, r_summary = _get_report_data(report_type, start_date_str, end_date_str, request=request)
        report_data = {
            'title': r_title, 'headers': r_headers, 'data': r_data, 'summary': r_summary
        }

    context = {
        'sales_metrics': sales_metrics,
        'top_products': top_products,
        'stock_metrics': stock_metrics,
        'low_stock_count': low_stock_count,
        # Flags para exibir as seções no template
        'show_sales': True,
        'show_products': True,
        'show_stock': True,
        'show_financial': True,
        'show_customers': True,
        'show_employees': True,
        # Listas vazias para evitar erros nos gráficos enquanto não implementamos a lógica complexa
        'monthly_profit': monthly_profit, # Agora populado
        'payment_stats': payment_stats,   # Novo dado
        'olfactory_stats': [],
        'peak_hours': peak_hours,         # Agora populado
        'monthly_expenses': [],
        'stock_by_brand': [],
        'dead_stock': [],
        'expiring_soon': low_stock_products, # Usando para mostrar estoque baixo
        'recent_movements': [],
        'recent_sales': recent_sales, # Nova lista de vendas
        'start_date': start_date_str,
        'end_date': end_date_str,
        'report_data': report_data,   # Dados do relatório visualizado
        'report_type': report_type,   # Tipo selecionado
        'period': period,
        'compare_start_date': compare_start_date_str,
        'compare_end_date': compare_end_date_str,
    }
    return render(request, 'reports/index.html', context)

@admin_required
@login_required
def download_report_file(request):
    report_type = request.GET.get('report_type', 'sales')
    file_format = request.GET.get('format', 'excel')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    title, headers, data, summary = _get_report_data(report_type, start_date, end_date, request=request)
    
    if file_format == 'excel':
        try:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{escape_uri_path("relatorio_" + report_type)}.xlsx"'
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório"
            
            ws.append([title])
            ws.append([f"Período: {start_date or 'Início'} a {end_date or 'Hoje'}"])
            ws.append([])
            ws.append(headers)
            
            for row in data:
                # Garante que os dados sejam strings para evitar erros de tipo no Excel
                ws.append([str(cell) if cell is not None else "" for cell in row])
                
            ws.append([])
            for k, v in summary.items():
                ws.append([k.replace('_', ' ').title(), str(v)])
                
            wb.save(response)
            return response
        except Exception as e:
            return HttpResponse(f"Erro ao gerar Excel: {str(e)}", status=500)

    elif file_format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path("relatorio_" + report_type)}.pdf"'
        
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Paragraph(f"Período: {start_date or 'Início'} a {end_date or 'Hoje'}", styles['Normal']))
        elements.append(Spacer(1, 12))
        
        table_data = [headers] + data
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(t)
        
        elements.append(Spacer(1, 12))
        for k, v in summary.items():
            elements.append(Paragraph(f"<b>{k.replace('_', ' ').title()}:</b> {v}", styles['Normal']))
            
        doc.build(elements)
        return response

    return redirect('reports_dashboard')

def _get_report_data(report_type, start_date, end_date, request=None):
    """
    Motor de Relatórios: Função genérica que retorna (Título, Cabeçalhos, Dados, Resumo).
    Isso permite que a mesma lógica alimente a visualização na tela, o PDF e o Excel,
    mantendo o princípio DRY (Don't Repeat Yourself).
    """
    data = []
    headers = []
    title = ""
    summary = {}

    # CORREÇÃO DO ERRO: Tratamento para datas "None" ou vazias
    if start_date in ['None', '', None]: start_date = None
    if end_date in ['None', '', None]: end_date = None

    if report_type == 'sales':
        qs = Sale.objects.all().select_related('customer').order_by('-created_at')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        title = "Relatório Geral de Vendas"
        headers = ['ID', 'Data', 'Cliente', 'Status', 'Pagamento', 'Total']
        for s in qs:
            data.append([s.id, s.created_at.strftime('%d/%m/%Y %H:%M'), s.customer.name if s.customer else 'Consumidor Final', s.get_status_display(), s.get_payment_method_display(), f"R$ {s.total:.2f}"])
        summary['total_vendas'] = f"R$ {sum(s.total for s in qs):.2f}"

    elif report_type == 'pending':
        qs = Sale.objects.filter(status='pending').select_related('customer').prefetch_related('items__product').order_by('-created_at')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        title = "Relatório de Vendas Pendentes / Orçamentos"
        headers = ['ID', 'Data', 'Cliente', 'Produtos', 'Qtd. Total', 'Valor Total']
        for s in qs:
            products_list = [f"{item.product.name} ({item.quantity})" for item in s.items.all()]
            products_str = "; ".join(products_list)
            total_qty = sum(item.quantity for item in s.items.all())
            data.append([s.id, s.created_at.strftime('%d/%m/%Y %H:%M'), s.customer.name if s.customer else 'Consumidor Final', products_str, total_qty, f"R$ {s.total:.2f}"])
        summary['total_pendente'] = f"R$ {sum(s.total for s in qs):.2f}"

    elif report_type == 'inventory':
        qs = Product.objects.all().order_by('name')
        
        # --- Filtros Profissionais Adicionais ---
        if request:
            brand_id = request.GET.get('brand')
            if brand_id and brand_id != 'all':
                qs = qs.filter(brand_id=brand_id)

            category_id = request.GET.get('category')
            if category_id and category_id != 'all':
                qs = qs.filter(category_id=category_id)

            supplier_id = request.GET.get('supplier')
            if supplier_id and supplier_id != 'all':
                qs = qs.filter(supplier_id=supplier_id)
                
            stock_status = request.GET.get('stock_status')
            if stock_status == 'in_stock':
                qs = qs.filter(stock_quantity__gt=0)
            elif stock_status == 'low_stock':
                qs = qs.filter(stock_quantity__lte=F('min_stock'))
            elif stock_status == 'out_of_stock':
                qs = qs.filter(stock_quantity__lte=0)

        title = "Relatório de Estoque e Valoração"
        headers = ['Produto', 'Marca', 'Estoque', 'Custo Unit.', 'Venda Unit.', 'Total Custo', 'Total Venda']
        total_cost = 0
        total_sale = 0
        for p in qs:
            t_cost = p.stock_quantity * p.cost_price
            t_sale = p.stock_quantity * p.selling_price
            total_cost += t_cost
            total_sale += t_sale
            data.append([p.name, p.brand.name if p.brand else '-', p.stock_quantity, f"R$ {p.cost_price:.2f}", f"R$ {p.selling_price:.2f}", f"R$ {t_cost:.2f}", f"R$ {t_sale:.2f}"])
        summary['custo_total'] = f"R$ {total_cost:.2f}"
        summary['venda_total'] = f"R$ {total_sale:.2f}"
        summary['lucro_potencial'] = f"R$ {total_sale - total_cost:.2f}"

    elif report_type == 'best_sellers':
        qs = SaleItem.objects.filter(sale__status='completed')
        if start_date: qs = qs.filter(sale__created_at__date__gte=start_date)
        if end_date: qs = qs.filter(sale__created_at__date__lte=end_date)
        qs = qs.values('product__name').annotate(total_qty=Sum('quantity'), total_rev=Sum(F('quantity') * F('price'))).order_by('-total_qty')
        title = "Produtos Mais Vendidos"
        headers = ['Produto', 'Qtd. Vendida', 'Receita Total']
        for item in qs:
            data.append([item['product__name'], item['total_qty'], f"R$ {item['total_rev']:.2f}"])

    elif report_type == 'sales_by_customer':
        qs = Sale.objects.filter(status='completed')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        qs = qs.values('customer__name').annotate(total_spent=Sum('total'), count=Count('id')).order_by('-total_spent')
        title = "Vendas por Cliente"
        headers = ['Cliente', 'Qtd. Compras', 'Total Gasto']
        for item in qs:
            name = item['customer__name'] if item['customer__name'] else 'Consumidor Final'
            data.append([name, item['count'], f"R$ {item['total_spent']:.2f}"])

    elif report_type == 'sales_by_brand':
        qs = SaleItem.objects.filter(sale__status='completed')
        if start_date: qs = qs.filter(sale__created_at__date__gte=start_date)
        if end_date: qs = qs.filter(sale__created_at__date__lte=end_date)
        qs = qs.values('product__brand__name').annotate(total_sold=Sum(F('quantity') * F('price')), qty=Sum('quantity')).order_by('-total_sold')
        title = "Vendas por Marca"
        headers = ['Marca', 'Qtd. Itens', 'Total Vendido']
        for item in qs:
            brand = item['product__brand__name'] if item['product__brand__name'] else 'Sem Marca'
            data.append([brand, item['qty'], f"R$ {item['total_sold']:.2f}"])

    elif report_type == 'sales_by_user':
        qs = Sale.objects.filter(status='completed')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        qs = qs.values('salesperson__username').annotate(total_sold=Sum('total'), count=Count('id')).order_by('-total_sold')
        title = "Vendas por Vendedor"
        headers = ['Vendedor', 'Qtd. Vendas', 'Total Vendido']
        for item in qs:
            user = item['salesperson__username'] if item['salesperson__username'] else 'Sistema'
            data.append([user, item['count'], f"R$ {item['total_sold']:.2f}"])

    elif report_type == 'sales_by_payment':
        qs = Sale.objects.filter(status='completed')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        qs = qs.values('payment_method').annotate(total=Sum('total'), count=Count('id')).order_by('-total')
        title = "Vendas por Forma de Pagamento"
        headers = ['Forma de Pagamento', 'Qtd. Vendas', 'Total']
        payment_map = {'pix': 'PIX', 'credit': 'Crédito', 'debit': 'Débito', 'cash': 'Dinheiro'}
        for item in qs:
            method = payment_map.get(item['payment_method'], item['payment_method'])
            data.append([method, item['count'], f"R$ {item['total']:.2f}"])

    elif report_type == 'profit_by_product':
        qs = SaleItem.objects.filter(sale__status='completed').select_related('product')
        if start_date: qs = qs.filter(sale__created_at__date__gte=start_date)
        if end_date: qs = qs.filter(sale__created_at__date__lte=end_date)
        
        # Agrupamento em Python pois o custo está no produto
        product_stats = {}
        for item in qs:
            pid = item.product.id
            if pid not in product_stats:
                product_stats[pid] = {'name': item.product.name, 'qty': 0, 'revenue': 0, 'cost': 0}
            
            product_stats[pid]['qty'] += item.quantity
            product_stats[pid]['revenue'] += item.price * item.quantity
            # Nota: Usa o custo atual do produto
            product_stats[pid]['cost'] += item.product.cost_price * item.quantity
            
        title = "Relatório de Lucro por Produto"
        headers = ['Produto', 'Qtd Vendida', 'Receita Total', 'Custo Total', 'Lucro Bruto', 'Margem %']
        
        total_gross_profit = 0
        sorted_stats = sorted(product_stats.values(), key=lambda x: x['revenue'] - x['cost'], reverse=True)
        
        for p in sorted_stats:
            profit = p['revenue'] - p['cost']
            margin = (profit / p['revenue'] * 100) if p['revenue'] > 0 else 0
            total_gross_profit += profit
            data.append([p['name'], p['qty'], f"R$ {p['revenue']:.2f}", f"R$ {p['cost']:.2f}", f"R$ {profit:.2f}", f"{margin:.1f}%"])
            
        summary['lucro_bruto_total'] = f"R$ {total_gross_profit:.2f}"

    elif report_type == 'profit_by_sale':
        qs = Sale.objects.filter(status='completed').prefetch_related('items__product').order_by('-created_at')
        if start_date: qs = qs.filter(created_at__date__gte=start_date)
        if end_date: qs = qs.filter(created_at__date__lte=end_date)
        
        title = "Relatório de Lucro por Venda"
        headers = ['ID Venda', 'Data', 'Cliente', 'Total Venda', 'Custo Produtos', 'Lucro', 'Margem %']
        
        total_profit_period = 0
        for s in qs:
            cost = sum(item.product.cost_price * item.quantity for item in s.items.all())
            profit = s.total - cost
            margin = (profit / s.total * 100) if s.total > 0 else 0
            total_profit_period += profit
            data.append([s.id, s.created_at.strftime('%d/%m/%Y'), s.customer.name if s.customer else 'Consumidor', f"R$ {s.total:.2f}", f"R$ {cost:.2f}", f"R$ {profit:.2f}", f"{margin:.1f}%"])
            
        summary['lucro_liquido_periodo'] = f"R$ {total_profit_period:.2f}"

    elif report_type == 'financial_expenses':
        qs = Expense.objects.all().order_by('-date')
        if start_date: qs = qs.filter(date__gte=start_date)
        if end_date: qs = qs.filter(date__lte=end_date)
        
        title = "Relatório de Despesas Operacionais"
        headers = ['Data', 'Descrição', 'Categoria', 'Valor']
        
        total_exp = 0
        for e in qs:
            total_exp += e.amount
            data.append([e.date.strftime('%d/%m/%Y'), e.description, e.get_category_display(), f"R$ {e.amount:.2f}"])
        
        summary['total_despesas'] = f"R$ {total_exp:.2f}"

    return title, headers, data, summary

@admin_required
@login_required
def export_dashboard(request):
    context = {
        'brands': Brand.objects.all().order_by('name'),
        'categories': Category.objects.all().order_by('name'),
        'suppliers': Supplier.objects.all().order_by('name'),
    }
    return render(request, 'reports/export.html', context)

@login_required
def export_data(request, model_name, file_format):
    # 1. Preparação dos Dados (Queryset e Headers)
    if model_name == 'products':
        queryset = Product.objects.all().select_related('brand', 'olfactory_family', 'category', 'supplier')
        
        # --- Filtros Profissionais de Exportação ---
        brand_id = request.GET.get('brand')
        if brand_id and brand_id != 'all':
            queryset = queryset.filter(brand_id=brand_id)

        category_id = request.GET.get('category')
        if category_id and category_id != 'all':
            queryset = queryset.filter(category_id=category_id)

        supplier_id = request.GET.get('supplier')
        if supplier_id and supplier_id != 'all':
            queryset = queryset.filter(supplier_id=supplier_id)
            
        stock_status = request.GET.get('stock_status')
        if stock_status == 'in_stock':
            queryset = queryset.filter(stock_quantity__gt=0)
        elif stock_status == 'low_stock':
            queryset = queryset.filter(stock_quantity__lte=F('min_stock'))
        elif stock_status == 'out_of_stock':
            queryset = queryset.filter(stock_quantity__lte=0)

        filename = 'produtos'
        headers = [
            'ID', 'Nome', 'Marca', 'Categoria', 'Fornecedor', 'Linha', 'Tipo', 'Gênero', 'Família Olfativa',
            'Notas de Saída', 'Notas de Corpo', 'Notas de Fundo', 'Descrição',
            'Volume', 'Código de Barras', 'Lote', 'Validade', 'Preço de Custo',
            'Preço de Venda', 'Qtd. em Estoque', 'Estoque Mínimo', 'URL da Imagem'
        ]
        def get_row(obj):
            return [
                str(obj.id),
                obj.name,
                obj.brand.name if obj.brand else '',
                obj.category.name if obj.category else '',
                obj.supplier.name if obj.supplier else '',
                obj.line,
                obj.get_product_type_display(),
                obj.get_gender_display(),
                obj.olfactory_family.name if obj.olfactory_family else '',
                obj.top_notes,
                obj.heart_notes,
                obj.base_notes,
                obj.description,
                obj.volume,
                obj.barcode,
                obj.batch_code,
                obj.expiration_date.strftime('%Y-%m-%d') if obj.expiration_date else '',
                f"{obj.cost_price:.2f}".replace('.', ','),
                f"{obj.selling_price:.2f}".replace('.', ','),
                str(obj.stock_quantity),
                str(obj.min_stock),
                obj.image_url if obj.image_url else ''
            ]
            
    elif model_name == 'sales':
        queryset = Sale.objects.all()
        filename = 'vendas'
        headers = ['ID', 'Data', 'Total', 'Status', 'Pagamento']
        def get_row(obj):
            return [str(obj.id), obj.created_at.strftime('%d/%m/%Y %H:%M'), str(obj.total), obj.get_status_display(), obj.payment_method]
            
    elif model_name == 'customers':
        queryset = Customer.objects.all()
        filename = 'clientes'
        headers = ['ID', 'Nome', 'Email', 'Telefone']
        def get_row(obj):
            return [str(obj.id), obj.name, obj.email, obj.phone]
            
    elif model_name == 'financial':
        queryset = Sale.objects.filter(status='completed').prefetch_related('items__product')
        filename = 'relatorio_financeiro_lucros'
        headers = ['ID Venda', 'Data', 'Total Venda', 'Custo Produtos', 'Lucro Bruto', 'Margem %']
        def get_row(obj):
            cost = sum(item.product.cost_price * item.quantity for item in obj.items.all())
            revenue = obj.total
            profit = revenue - cost
            margin = (profit / revenue * 100) if revenue > 0 else 0
            return [str(obj.id), obj.created_at.strftime('%d/%m/%Y %H:%M'), f"{revenue:.2f}", f"{cost:.2f}", f"{profit:.2f}", f"{margin:.2f}"]
    else:
        return JsonResponse({'status': 'error', 'message': 'Modelo inválido'})

    # 2. Geração da Lista de Dados
    data_rows = []
    for obj in queryset:
        data_rows.append(get_row(obj))

    # 3. Renderização por Formato
    if file_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(data_rows)
        return response

    elif file_format == 'excel':
        if not openpyxl:
            return HttpResponse("Erro: Biblioteca 'openpyxl' não instalada. Instale com: pip install openpyxl", status=500)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados"

        # Cabeçalho Profissional no Excel
        ws.append([f"Relatório de {model_name.title()} - Perfume ERP"])
        ws.append([f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append([])
        ws.append(headers)
        for row in data_rows:
            ws.append(row)
        wb.save(response)
        return response

    elif file_format == 'pdf':
        if not canvas:
            return HttpResponse("Erro: Biblioteca 'reportlab' não instalada. Instale com: pip install reportlab", status=500)

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image as PDFImage, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph(f"Relatório: {filename.replace('_', ' ').title()}", styles['Title']))
        
        # Lógica personalizada para Produtos (com Imagem e Margem)
        if model_name == 'products':
            # Layout em Grade (Catálogo Visual)
            # Recria estilos para não afetar o global
            cat_styles = getSampleStyleSheet()
            style_center = cat_styles['Normal']
            style_center.alignment = 1 # Center
            style_center.fontSize = 9
            
            style_title = cat_styles['Normal']
            style_title.alignment = 1
            style_title.fontName = 'Helvetica-Bold'
            style_title.fontSize = 10
            
            style_price = cat_styles['Normal']
            style_price.alignment = 1
            style_price.fontName = 'Helvetica-Bold'
            style_price.fontSize = 12
            style_price.textColor = colors.darkgreen

            # Configuração do Grid (4 colunas para Landscape)
            cols = 4
            col_width = 65 * mm
            data_matrix = []
            current_row = []

            for obj in queryset:
                # Imagem
                img_obj = ""
                if obj.image:
                    try:
                        if os.path.exists(obj.image.path):
                            # Imagem maior para catálogo (40x40mm)
                            img_obj = PDFImage(obj.image.path, width=40*mm, height=40*mm)
                        else:
                            img_obj = Paragraph("Imagem não encontrada", style_center)
                    except Exception:
                        img_obj = Paragraph("Erro", style_center)
                else:
                    img_obj = Paragraph("Sem Foto", style_center)

                # Monta o Card do Produto (Célula)
                cell_content = [
                    img_obj,
                    Spacer(1, 2*mm),
                    Paragraph(obj.name, style_title),
                    Paragraph(obj.brand.name if obj.brand else "-", style_center),
                    Spacer(1, 1*mm),
                    Paragraph(f"R$ {obj.selling_price:.2f}", style_price),
                    Paragraph(f"Estoque: {obj.stock_quantity}", style_center)
                ]
                current_row.append(cell_content)
                
                if len(current_row) >= cols:
                    data_matrix.append(current_row)
                    current_row = []
            
            # Completa a última linha
            if current_row:
                while len(current_row) < cols:
                    current_row.append("")
                data_matrix.append(current_row)

            if data_matrix:
                t = Table(data_matrix, colWidths=[col_width]*cols)
                t.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("Nenhum produto cadastrado.", styles['Normal']))
            
        else:
            # Lógica Genérica para outros modelos (Vendas, Clientes, etc)
            table_data = [headers] + data_rows
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
        elements.append(t)
        doc.build(elements)
        return response
        
    elif file_format == 'json':
        json_data = [dict(zip(headers, row)) for row in data_rows]
        return JsonResponse(json_data, safe=False, encoder=DjangoJSONEncoder)

    return JsonResponse({'status': 'error', 'message': 'Formato não suportado'})

@admin_required
@login_required
def import_data(request):
    # --- Lógica de Importação de Arquivos ---
    if request.method == 'POST' and request.FILES.get('file'):
        uploaded_file = request.FILES['file']
        filename = uploaded_file.name.lower()
        model_type = request.POST.get('model') # Captura o tipo selecionado (se houver)
        
        try:
            data_list = []
            
            # 1. Parse File
            if filename.endswith('.csv'):
                # Wrapper para ler bytes como texto (utf-8-sig lida com BOM do Excel)
                # Tenta UTF-8 primeiro, se falhar, tenta Latin-1 (comum no Excel BR)
                try:
                    file_file = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig')
                    sample = file_file.read(2048)
                    file_file.seek(0)
                except UnicodeDecodeError:
                    uploaded_file.file.seek(0)
                    file_file = TextIOWrapper(uploaded_file.file, encoding='latin-1')
                    sample = file_file.read(2048)
                    file_file.seek(0)

                # Tenta detectar se usa ponto e vírgula (comum no Excel Brasil) ou vírgula
                delimiter = ';' if sample.count(';') > sample.count(',') else ','
                reader = csv.DictReader(file_file, delimiter=delimiter)
                data_list = list(reader)
            elif filename.endswith('.json'):
                data_list = json.load(uploaded_file)
            elif filename.endswith('.xlsx'):
                if not openpyxl:
                    messages.error(request, 'Biblioteca openpyxl não instalada no servidor.')
                    return redirect('import_data')
                wb = openpyxl.load_workbook(uploaded_file)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    # 1. Busca inteligente da linha de cabeçalho (ignora linhas vazias ou títulos no topo)
                    header_row_index = 0
                    for i, row in enumerate(rows[:20]): # Verifica as primeiras 20 linhas
                        row_values = [str(cell).lower().strip() for cell in row if cell is not None]
                        # Palavras-chave para identificar se é um cabeçalho válido
                        keywords = ['nome', 'name', 'código', 'codigo', 'barcode', 'preço', 'price', 'estoque', 'stock', 'email', 'telefone']
                        if any(k in row_values for k in keywords):
                            header_row_index = i
                            break
                    
                    # 2. Mapeia as colunas baseadas no cabeçalho encontrado
                    raw_headers = rows[header_row_index]
                    header_map = {i: str(h).lower().strip() for i, h in enumerate(raw_headers) if h is not None and str(h).strip()}
                    
                    # 3. Processa os dados
                    for row in rows[header_row_index + 1:]:
                        row_dict = {}
                        has_data = False
                        for i, cell in enumerate(row):
                            if i in header_map:
                                val = str(cell) if cell is not None else ''
                                # Remove .0 de números inteiros (comum no Excel)
                                if val.endswith('.0'): val = val[:-2]
                                row_dict[header_map[i]] = val.strip()
                                if val.strip(): has_data = True
                        if has_data:
                            data_list.append(row_dict)
            else:
                messages.error(request, 'Formato inválido. Use .csv, .json ou .xlsx')
                return redirect('import_data')
            
            if not data_list:
                messages.warning(request, 'O arquivo está vazio.')
                return redirect('import_data')

            # 2. Identify Model based on headers (heuristic)
            first_row = {k.lower().strip(): v for k, v in data_list[0].items()}
            keys = first_row.keys()
            
            success_count = 0
            errors = []
            model_name = ""

            # Helper para limpar valores monetários (Ex: "R$ 1.200,50" -> 1200.50)
            def clean_decimal(val):
                if not val: return None
                s = str(val).replace('R$', '').replace(' ', '').strip()
                # Se tiver ponto e vírgula, assume formato BR (milhar.centena,decimal)
                if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
                elif ',' in s: s = s.replace(',', '.')
                try: return Decimal(s)
                except: return Decimal('0')

            # Lógica para Produtos (busca por colunas típicas)
            # Verifica se foi selecionado 'products' OU se os cabeçalhos indicam produtos
            if model_type == 'products' or (not model_type and any(k in keys for k in ['estoque', 'stock', 'stock_quantity', 'preço', 'price', 'selling_price'])):
                model_name = "Produtos"
                for row in data_list:
                    try:
                        row_lower = {k.lower().strip(): v for k, v in row.items()}
                        
                        pk = row_lower.get('id')
                        barcode = row_lower.get('código de barras') or row_lower.get('codigo de barras') or row_lower.get('barcode')
                        
                        if barcode:
                            barcode = str(barcode).strip()
                            if not barcode: barcode = None
                        else:
                            barcode = None

                        name = row_lower.get('nome') or row_lower.get('name')
                        
                        if not name:
                            errors.append(f"Linha ignorada por não conter 'Nome': {row}")
                            continue

                        # --- Lógica de Relacionamentos (Foreign Keys) ---
                        brand_obj = None
                        brand_name = row_lower.get('marca') or row_lower.get('brand')
                        if brand_name:
                            brand_obj, _ = Brand.objects.get_or_create(name__iexact=brand_name.strip(), defaults={'name': brand_name.strip()})

                        olfactory_family_obj = None
                        olfactory_family_name = row_lower.get('família olfativa') or row_lower.get('familia olfativa')
                        if olfactory_family_name:
                            olfactory_family_obj, _ = OlfactoryFamily.objects.get_or_create(name__iexact=olfactory_family_name.strip(), defaults={'name': olfactory_family_name.strip()})
                        
                        # --- Busca o produto existente ---
                        product = None
                        if pk and Product.objects.filter(pk=pk).exists():
                            product = Product.objects.get(pk=pk)
                        elif barcode and Product.objects.filter(barcode=barcode).exists():
                            product = Product.objects.get(barcode=barcode)
                        elif name:
                            # Busca por nome e marca para ser mais preciso
                            search_filter = {'name__iexact': name}
                            if brand_obj:
                                search_filter['brand'] = brand_obj
                            product = Product.objects.filter(**search_filter).first()
                        
                        # --- Prepara os dados ---
                        product_data = {
                            'name': name,
                            'line': row_lower.get('linha'),
                            'product_type': row_lower.get('tipo'),
                            'gender': (row_lower.get('gênero') or row_lower.get('genero') or '').upper(),
                            'olfactory_family': olfactory_family_obj,
                            'top_notes': row_lower.get('notas de saída'),
                            'heart_notes': row_lower.get('notas de corpo'),
                            'base_notes': row_lower.get('notas de fundo'),
                            'description': row_lower.get('descrição') or row_lower.get('descricao'),
                            'volume': row_lower.get('volume'),
                            'barcode': barcode,
                            'batch_code': row_lower.get('lote'),
                            'expiration_date': row_lower.get('validade') or None,
                            'cost_price': clean_decimal(row_lower.get('preço de custo') or row_lower.get('preco de custo') or row_lower.get('custo')),
                            'selling_price': clean_decimal(row_lower.get('preço de venda') or row_lower.get('preco de venda') or row_lower.get('preço') or row_lower.get('price')),
                            'stock_quantity': row_lower.get('qtd. em estoque') or row_lower.get('qtd em estoque') or row_lower.get('estoque'),
                            'min_stock': row_lower.get('estoque mínimo') or row_lower.get('estoque minimo'),
                            'image_url': row_lower.get('url da imagem') or row_lower.get('imagem')
                        }
                        if brand_obj:
                            product_data['brand'] = brand_obj

                        # --- Limpa e Converte os Tipos ---
                        # Já convertido pelo clean_decimal acima
                        if product_data['stock_quantity']: product_data['stock_quantity'] = int(float(str(product_data['stock_quantity']).replace(',', '.')))
                        if product_data['min_stock']: product_data['min_stock'] = int(float(str(product_data['min_stock']).replace(',', '.')))
                        if not product_data['expiration_date']: del product_data['expiration_date']

                        # Filtra valores nulos/vazios para não sobrescrever dados existentes com nada
                        update_data = {k: v for k, v in product_data.items() if v is not None and v != ''}
                        
                        if product: # Atualiza
                            for key, value in update_data.items():
                                setattr(product, key, value)
                            product.save()
                            success_count += 1
                        elif name and update_data.get('selling_price'): # Cria
                            # Garante campos obrigatórios para criação
                            if 'brand' not in update_data:
                                default_brand, _ = Brand.objects.get_or_create(name="Geral")
                                update_data['brand'] = default_brand
                            if 'cost_price' not in update_data:
                                update_data['cost_price'] = 0
                            if 'volume' not in update_data:
                                update_data['volume'] = 'N/A'
                            
                            Product.objects.create(**update_data)
                            success_count += 1
                    except Exception as e:
                        errors.append(f"Erro na linha {row}: {str(e)}")

            # Lógica para Clientes
            # Verifica se foi selecionado 'customers' OU se os cabeçalhos indicam clientes
            elif model_type == 'customers' or (not model_type and any(k in keys for k in ['email', 'telefone', 'phone'])):
                model_name = "Clientes"
                for row in data_list:
                    try:
                        row_lower = {k.lower().strip(): v for k, v in row.items()}
                        pk = row_lower.get('id')
                        defaults = {
                            'name': row_lower.get('nome') or row_lower.get('name'),
                            'email': row_lower.get('email'),
                            'phone': row_lower.get('telefone') or row_lower.get('phone')
                        }
                        # Remove chaves vazias
                        defaults = {k: v for k, v in defaults.items() if v}
                        
                        if pk:
                            Customer.objects.update_or_create(pk=pk, defaults=defaults)
                        elif defaults.get('email'):
                            Customer.objects.update_or_create(email=defaults['email'], defaults=defaults)
                        elif defaults.get('name'):
                            Customer.objects.update_or_create(name=defaults['name'], defaults=defaults)
                        success_count += 1
                    except Exception as e:
                        errors.append(f"Erro na linha {row}: {str(e)}")
            
            else:
                messages.error(request, 'Não foi possível identificar o tipo de dados (Produtos ou Clientes). Verifique os cabeçalhos.')
                return redirect('import_data')

            if success_count > 0:
                messages.success(request, f'{success_count} registros de {model_name} processados com sucesso!')
            
            if errors:
                for err in errors[:3]: messages.warning(request, err)
                if len(errors) > 3: messages.warning(request, f'E mais {len(errors)-3} erros.')

        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')

        return redirect('import_data')

    # --- Lógica Fiscal: Listar Vendas Recentes ---
    recent_sales = Sale.objects.filter(status__in=['completed', 'pending']).select_related('customer').order_by('-created_at')[:10]

    return render(request, 'reports/import.html', {'sales': recent_sales})

@admin_required
@login_required
def download_fiscal(request, sale_id, doc_type):
    sale = get_object_or_404(Sale, pk=sale_id)
    
    # Busca dados da empresa
    try:
        company = CompanySettings.objects.first()
    except (OperationalError, ProgrammingError):
        company = None
    company_name = company.name if company else "Perfume ERP Ltda"
    company_cnpj = company.cnpj if company else "00.000.000/0001-91"
    
    if doc_type == 'cupom':
        if not canvas:
            messages.error(request, 'PDFs indisponíveis (ReportLab ausente).')
            return redirect('import_data')
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cupom_{sale.id}.pdf"'
        
        # Gera PDF estilo Cupom Térmico (80mm largura) - Layout Moderno
        width = 80*mm
        height = 250*mm # Altura maior para garantir que caibam os itens
        p = canvas.Canvas(response, pagesize=(width, height))
        
        # Coordenadas e Configurações
        y = 240*mm 
        left_x = 5*mm
        center_x = 40*mm
        right_x = 75*mm
        line_height = 4*mm
        
        # --- Cabeçalho da Empresa ---
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(center_x, y, company_name.upper())
        y -= line_height + 1*mm
        
        p.setFont("Helvetica", 8)
        p.drawCentredString(center_x, y, f"CNPJ: {company_cnpj}")
        y -= line_height
        
        if company:
            if company.address:
                p.drawCentredString(center_x, y, company.address[:45]) # Trunca para caber
                y -= line_height
            if company.phone:
                p.drawCentredString(center_x, y, f"Tel: {company.phone}")
                y -= line_height
        
        y -= 2*mm
        p.setLineWidth(0.5)
        p.line(left_x, y, right_x, y)
        y -= line_height + 2*mm
        
        # --- Dados da Venda ---
        p.setFont("Helvetica-Bold", 9)
        p.drawCentredString(center_x, y, "CUPOM NÃO FISCAL")
        y -= line_height
        
        p.setFont("Helvetica", 8)
        p.drawCentredString(center_x, y, f"Venda Nº {sale.id:06d}")
        y -= line_height
        p.drawCentredString(center_x, y, sale.created_at.strftime('%d/%m/%Y %H:%M:%S'))
        y -= line_height + 2*mm
        
        p.line(left_x, y, right_x, y)
        y -= line_height + 2*mm
        
        # --- Itens ---
        p.setFont("Helvetica-Bold", 8)
        p.drawString(left_x, y, "ITEM")
        p.drawRightString(right_x, y, "TOTAL")
        y -= line_height
        
        p.setFont("Helvetica", 8)
        for item in sale.items.all():
            # Nome do Produto
            name = item.product.name
            if len(name) > 30: name = name[:30] + "..."
            p.drawString(left_x, y, name)
            y -= line_height
            
            # Detalhes
            details = f"{item.quantity} x R$ {item.price:.2f}"
            total_item = item.quantity * item.price
            
            p.setFont("Helvetica", 7)
            p.drawString(left_x + 2*mm, y, details)
            p.drawRightString(right_x, y, f"R$ {total_item:.2f}")
            p.setFont("Helvetica", 8)
            y -= line_height + 1*mm
            
        y -= 2*mm
        p.line(left_x, y, right_x, y)
        y -= line_height + 2*mm
        
        # --- Totais ---
        p.setFont("Helvetica-Bold", 12)
        p.drawString(left_x, y, "TOTAL A PAGAR")
        p.drawRightString(right_x, y, f"R$ {sale.total:.2f}")
        y -= line_height + 4*mm
        
        p.setFont("Helvetica", 8)
        p.drawString(left_x, y, "Forma de Pagamento:")
        p.drawRightString(right_x, y, sale.get_payment_method_display())
        y -= line_height + 2*mm
        
        # --- Cliente ---
        if sale.customer:
            p.line(left_x, y, right_x, y)
            y -= line_height + 2*mm
            p.setFont("Helvetica-Bold", 8)
            p.drawString(left_x, y, "CLIENTE")
            y -= line_height
            p.setFont("Helvetica", 8)
            p.drawString(left_x, y, sale.customer.name[:35])
            y -= line_height
            if sale.customer.cpf_cnpj:
                p.drawString(left_x, y, f"CPF/CNPJ: {sale.customer.cpf_cnpj}")
                y -= line_height
        
        # --- Rodapé ---
        y -= 10*mm
        p.setFont("Helvetica-Oblique", 8)
        p.drawCentredString(center_x, y, "Obrigado pela preferência!")
        y -= line_height
        p.setFont("Helvetica", 6)
        p.drawCentredString(center_x, y, "Gerado por Perfume ERP")

        # --- QR Code ---
        y -= 5*mm
        qr_data = "https://www.seusite.com.br" # Substitua pelo seu site
        qr_code = qr.QrCodeWidget(qr_data)
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(45, 45, transform=[45/width,0,0,45/height,0,0])
        d.add(qr_code)
        renderPDF.draw(d, p, center_x - 8*mm, y - 18*mm)

        p.showPage()
        p.save()
        return response

    elif doc_type == 'nfe':
        # Gera XML Simulado de NF-e
        response = HttpResponse(content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="nfe_{sale.id}.xml"'
        
        xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<nfeProc xmlns="http://www.portalfiscal.inf.br/nfe" versao="4.00">
    <NFe>
        <infNFe Id="NFe{sale.id}">
            <ide>
                <nNF>{sale.id}</nNF>
                <dhEmi>{sale.created_at.isoformat()}</dhEmi>
                <tpNF>1</tpNF>
                <natOp>Venda de Mercadoria</natOp>
            </ide>
            <emit>
                <xNome>{company_name}</xNome>
                <CNPJ>{company_cnpj}</CNPJ>
            </emit>
            <dest>
                <xNome>{sale.customer.name if sale.customer else 'Consumidor Final'}</xNome>
                <CPF>{sale.customer.cpf_cnpj if sale.customer and sale.customer.cpf_cnpj else ''}</CPF>
            </dest>
            <det nItem="1">
                <prod>
                    <xProd>Venda de Perfumes e Cosmeticos</xProd>
                    <vProd>{sale.total}</vProd>
                </prod>
            </det>
            <total>
                <ICMSTot>
                    <vNF>{sale.total}</vNF>
                </ICMSTot>
            </total>
        </infNFe>
    </NFe>
</nfeProc>"""
        response.write(xml_content)
        return response
        
    elif doc_type == 'danfe':
        if not canvas:
            messages.error(request, 'PDFs indisponíveis (ReportLab ausente).')
            return redirect('import_data')
        # Gera PDF estilo DANFE (Simplificado A4)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="danfe_{sale.id}.pdf"'
        
        # Configuração A4 (210mm x 297mm)
        width, height = 210*mm, 297*mm
        p = canvas.Canvas(response, pagesize=(width, height))
        
        # --- Cabeçalho ---
        p.setLineWidth(1)
        p.rect(10*mm, height - 40*mm, 190*mm, 30*mm)
        
        p.setFont("Helvetica-Bold", 12)
        p.drawString(15*mm, height - 20*mm, "DANFE - Documento Auxiliar da Nota Fiscal Eletrônica")
        
        p.setFont("Helvetica", 10)
        p.drawString(15*mm, height - 25*mm, company_name)
        p.drawString(15*mm, height - 30*mm, f"CNPJ: {company_cnpj}")
        
        p.drawString(120*mm, height - 20*mm, f"Nº: {sale.id}")
        p.drawString(120*mm, height - 25*mm, "Série: 1")
        p.drawString(120*mm, height - 30*mm, f"Emissão: {sale.created_at.strftime('%d/%m/%Y')}")
        
        # --- Destinatário ---
        y = height - 50*mm
        p.rect(10*mm, y - 25*mm, 190*mm, 25*mm)
        p.setFont("Helvetica-Bold", 10)
        p.drawString(12*mm, y - 5*mm, "DESTINATÁRIO / REMETENTE")
        
        p.setFont("Helvetica", 9)
        customer_name = sale.customer.name if sale.customer else 'Consumidor Final'
        customer_doc = sale.customer.cpf_cnpj if sale.customer and sale.customer.cpf_cnpj else ''
        
        p.drawString(15*mm, y - 12*mm, f"Nome/Razão Social: {customer_name}")
        p.drawString(130*mm, y - 12*mm, f"CNPJ/CPF: {customer_doc}")
        
        # --- Itens ---
        y = height - 85*mm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(10*mm, y, "DADOS DOS PRODUTOS / SERVIÇOS")
        y -= 5*mm
        p.line(10*mm, y, 200*mm, y)
        y -= 5*mm
        
        p.setFont("Helvetica", 9)
        for item in sale.items.all():
            line_text = f"{item.product.name[:50]} | Qtd: {item.quantity} | Unit: R$ {item.price:.2f} | Total: R$ {item.quantity * item.price:.2f}"
            p.drawString(12*mm, y, line_text)
            y -= 5*mm
            
        # --- Totais ---
        y -= 5*mm
        p.line(10*mm, y, 200*mm, y)
        y -= 8*mm
        p.setFont("Helvetica-Bold", 11)
        p.drawString(140*mm, y, f"VALOR TOTAL: R$ {sale.total:.2f}")
        
        p.showPage()
        p.save()
        return response

    return redirect('import_data')

@admin_required
@login_required
def download_db_backup(request):
    db_settings = settings.DATABASES['default']
    
    # Verifica se é SQLite
    if db_settings['ENGINE'] != 'django.db.backends.sqlite3':
        return JsonResponse({'status': 'error', 'message': 'Backup automático disponível apenas para SQLite'})
        
    db_path = db_settings['NAME']
    if os.path.exists(db_path):
        with open(db_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/x-sqlite3")
            response['Content-Disposition'] = 'attachment; filename="backup_db.sqlite3"'
            return response
            
    return JsonResponse({'status': 'error', 'message': 'Arquivo do banco de dados não encontrado'})

@login_required
def pending_sales(request):
    sales = Sale.objects.filter(status='pending').select_related('customer').prefetch_related('items__product').order_by('-created_at')
    
    for sale in sales:
        # Garante que amount_paid não seja None
        if sale.amount_paid is None:
            sale.amount_paid = Decimal('0')
        sale.remaining = sale.total - sale.amount_paid
        
    return render(request, 'reports/pending_sales.html', {'sales': sales})

@admin_required
@login_required
def delete_sale(request, sale_id):
    sale = get_object_or_404(Sale, pk=sale_id)
    
    # Estorno de Estoque (Seja pendente ou finalizada)
    # Devolve os itens ao estoque antes de excluir
    for item in sale.items.all():
        StockMovement.objects.create(
            product=item.product,
            quantity=item.quantity,
            movement_type='E',
            reason=f'Estorno/Exclusão Venda #{sale.id}'
        )
        
    sale.delete()
    messages.success(request, f'Venda #{sale_id} excluída e estoque estornado.')
    
    # Redireciona para a página anterior (Dashboard ou Pendentes)
    referer = request.META.get('HTTP_REFERER', '')
    if 'dashboard' in referer or 'detalhe' in referer:
        return redirect('reports_dashboard')
    return redirect('pending_sales')

@admin_required
@login_required
def delete_sale_item(request, item_id):
    item = get_object_or_404(SaleItem, pk=item_id)
    sale = item.sale
    
    try:
        with transaction.atomic():
            # 1. Devolve o item ao estoque
            StockMovement.objects.create(
                product=item.product,
                quantity=item.quantity,
                movement_type='E', # Entrada (estorno)
                reason=f'Remoção de item da Venda #{sale.id}'
            )
            
            # 2. Deleta o item
            item.delete()
            
            # 3. Verifica se a venda ficou vazia
            if not sale.items.exists():
                sale.delete()
                messages.success(request, 'Último item removido. A venda foi excluída e o estoque totalmente estornado.')
                return redirect('reports_dashboard')
            else:
                # Apenas recalcula o total se ainda houver itens
                sale.save()
                messages.success(request, 'Item removido da venda e estoque estornado.')

    except Exception as e:
        messages.error(request, f'Erro ao remover item: {str(e)}')
    
    return redirect('sale_detail', sale_id=sale.id)

@login_required
def sale_detail(request, sale_id):
    """
    Exibe detalhes de uma venda e permite edição simples (Cliente/Pagamento) ou exclusão.
    """
    sale = get_object_or_404(Sale.objects.prefetch_related('items__product'), pk=sale_id)
    customers = Customer.objects.all().order_by('name')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Atualizar Metadados (Cliente, Data, Pagamento, Descontos)
                customer_id = request.POST.get('customer')
                sale.customer_id = customer_id if customer_id else None
                
                sale.payment_method = request.POST.get('payment_method')
                sale.created_at = request.POST.get('created_at')
                

                sale.discount_value = clean_br_decimal(request.POST.get('discount_value', '0'))
                sale.discount_type = request.POST.get('discount_type', 'fixed')
                
                sale.tax_value = clean_br_decimal(request.POST.get('tax_value', '0'))
                sale.tax_type = request.POST.get('tax_type', 'fixed')

                # 2. Atualizar Itens (Quantidade e Preço)
                for item in sale.items.all():
                    qty_key = f'quantity_{item.id}'
                    price_key = f'price_{item.id}'
                    
                    if qty_key in request.POST and price_key in request.POST:
                        # Conversão segura para int (ex: "1.0" vira 1, "1" vira 1)
                        try:
                            new_qty = int(float(request.POST[qty_key]))
                        except ValueError:
                            new_qty = item.quantity

                        new_price = clean_br_decimal(request.POST[price_key])
                        
                        # Ajuste de Estoque pela diferença
                        diff = new_qty - item.quantity
                        if diff != 0:
                            prod = Product.objects.select_for_update().get(pk=item.product.id)
                            StockMovement.objects.create(
                                product=prod,
                                quantity=abs(diff),
                                movement_type='S' if diff > 0 else 'E', # S=Saída (aumentou venda), E=Entrada (diminuiu)
                                reason=f'Edição Venda #{sale.id}'
                            )
                        
                        item.quantity = new_qty
                        item.price = new_price
                        item.save(update_sale_total=False) # Atualiza subtotal do item, mas não salva a venda ainda

                # 3. Salvar Venda (Recalcula o TOTAL GERAL baseado nos itens e descontos novos)
                sale.save()
                messages.success(request, 'Venda atualizada com sucesso!')
                return redirect('sale_detail', sale_id=sale.id)
                
        except Exception as e:
            messages.error(request, f'Erro ao salvar: {str(e)}')
            
    subtotal_items = sum(item.subtotal for item in sale.items.all())
    return render(request, 'reports/sale_detail.html', {'sale': sale, 'customers': customers, 'subtotal_items': subtotal_items})

@admin_required
@login_required
def expense_manage(request):
    """
    View para Gerenciamento de Despesas.
    Permite adicionar despesas únicas ou PARCELADAS.
    """
    if request.method == 'POST':
        try:
            action = request.POST.get('action', 'add')
            
            if action == 'delete':
                # Lógica de exclusão
                pk = request.POST.get('expense_id')
                Expense.objects.filter(pk=pk).delete()
                messages.success(request, 'Despesa removida com sucesso.')
            
            elif action == 'edit':
                pk = request.POST.get('expense_id')
                exp = get_object_or_404(Expense, pk=pk)
                exp.description = request.POST.get('description')
                exp.category = request.POST.get('category')
                exp.amount = clean_br_decimal(request.POST.get('amount'))
                if request.POST.get('date'):
                    exp.date = request.POST.get('date')
                exp.paid = request.POST.get('paid') == 'on'
                exp.save()
                messages.success(request, 'Despesa atualizada com sucesso!')

            elif action == 'add':
                # Captura dados do formulário
                description = request.POST.get('description')
                category = request.POST.get('category')
                
                # Limpa formato de moeda (R$ 1.000,00 -> 1000.00)
                amount = clean_br_decimal(request.POST.get('amount', '0'))
                
                date_str = request.POST.get('date')
                start_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # Lógica de Parcelamento
                installments = int(request.POST.get('installments', 1))
                paid = request.POST.get('paid') == 'on'
                
                if installments > 1:
                    # Se for parcelado, divide o valor e cria múltiplos registros
                    installment_value = amount / installments
                    for i in range(installments):
                        # Lógica para adicionar meses corretamente
                        month = start_date.month - 1 + i
                        # Calcula ano e mês corretos (virada de ano)
                        year = start_date.year + month // 12
                        month = month % 12 + 1
                        # Ajusta dia para o último dia do mês se necessário (ex: 31/01 -> 28/02)
                        day = min(start_date.day, calendar.monthrange(year, month)[1])
                        due_date = date(year, month, day)
                        
                        Expense.objects.create(
                            description=f"{description} ({i+1}/{installments})",
                            category=category,
                            amount=installment_value,
                            date=due_date,
                            paid=paid if i == 0 else False # Apenas a primeira parcela conta como paga se marcado
                        )
                    messages.success(request, f'Despesa parcelada em {installments}x lançada com sucesso!')
                else:
                    Expense.objects.create(
                        # Despesa Simples (À vista)
                        description=description,
                        category=category,
                        amount=amount,
                        date=start_date,
                        paid=paid
                    )
                    messages.success(request, 'Despesa registrada!')
        except Exception as e:
            messages.error(request, f'Erro ao lançar despesa: {str(e)}')
        return redirect('expense_manage')

    # --- Nova Lógica de Agrupamento para Visualização ---
    raw_expenses = Expense.objects.all().order_by('-date')
    
    groups = {}
    for exp in raw_expenses:
        # Identifica padrão "Nome da Conta (1/3)" para agrupar
        match = re.search(r'^(.*)\s\(\d+/\d+\)$', exp.description)
        if match:
            group_name = match.group(1) # Pega apenas "Nome da Conta"
        else:
            group_name = exp.description
            
        if group_name not in groups:
            groups[group_name] = {
                'name': group_name,
                'items': [],
                'total': Decimal('0'),
                'last_date': exp.date, # Para ordenar pelo mais recente
                'has_installments': False
            }
        
        if match:
            groups[group_name]['has_installments'] = True
            
        groups[group_name]['items'].append(exp)
        groups[group_name]['total'] += exp.amount
        if exp.date > groups[group_name]['last_date']:
            groups[group_name]['last_date'] = exp.date

    # Ordena os grupos pela data mais recente
    grouped_list = sorted(groups.values(), key=lambda x: x['last_date'], reverse=True)
    total = raw_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    
    return render(request, 'reports/expenses.html', {
        'expenses': grouped_list, 
        'total': total,
        'category_choices': Expense.CATEGORY_CHOICES
    })

@login_required
def register_payment(request, sale_id):
    """
    Processa o pagamento de uma venda pendente (Parcial ou Total).
    """
    sale = get_object_or_404(Sale, pk=sale_id)
    if request.method == 'POST':
        try:
            action_type = request.POST.get('action_type') # 'save' ou 'finalize'
            payment_method = request.POST.get('payment_method')
            
            # Limpa o valor monetário vindo do formulário (ex: "1.234,56" -> "1234.56")
            val_str = request.POST.get('amount', '0')
            # Helper para garantir conversão correta BR (com milhar) vs US/Simples
            if ',' in val_str and '.' in val_str: # Ex: 1.234,56
                val_str = val_str.replace('.', '').replace(',', '.')
            elif ',' in val_str: # Ex: 1234,56
                val_str = val_str.replace(',', '.')
            
            payment_value = Decimal(val_str)

            with transaction.atomic():
                # Atualiza valor pago acumulado
                sale.amount_paid += payment_value
                
                # Se o usuário clicou em "Finalizar" OU o valor acumulado já cobre o total
                if action_type == 'finalize' or sale.amount_paid >= sale.total:
                    if action_type == 'finalize' and sale.amount_paid < sale.total:
                        # Se forçou finalizar mas valor era menor, assume que pagou o resto agora
                        sale.amount_paid = sale.total
                    
                    sale.status = 'completed'
                    sale.payment_method = payment_method
                    sale.save()
                    # Importante: Chama a finalização para baixar estoque corretamente
                    if hasattr(sale, 'finalize'):
                        sale.finalize()
                    messages.success(request, f'Venda #{sale.id} FINALIZADA com sucesso!')
                else:
                    # Apenas salvou parcial
                    sale.save()
                    remaining = sale.total - sale.amount_paid
                    messages.info(request, f'Pagamento parcial de R$ {payment_value:.2f} salvo. Restam R$ {remaining:.2f}.')

        except Exception as e:
            messages.error(request, f'Erro ao registrar pagamento: {str(e)}')
            
    return redirect('pending_sales')

@admin_required
@login_required
def audit_logs(request):
    """Exibe o histórico de logs do sistema"""
    logs = AuditLog.objects.select_related('user').all()[:500] # Limite de 500 para performance
    return render(request, 'reports/audit_logs.html', {'logs': logs})